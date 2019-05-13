
import sys
import openpyxl
import datetime
import traceback
import os
import zipfile

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment, Protection, colors
from openpyxl.comments import Comment
from openpyxl.worksheet	import Worksheet
from openpyxl import *
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import timedelta
from string import maketrans
from boxsdk import Client, OAuth2
from boxsdk.network.default_network import DefaultNetwork

from trnng_lst import trnng_lst_scnd

reload(sys)
sys.setdefaultencoding('utf8')

calendar = []
nme_lst = []

def border_line(work_sheet, row_start, row_end, column_start, column_end):
	#make border line in each cell
	
	try:
		for row_index in range(row_start, row_end):
			for col_index in range(column_start, column_end):
				work_sheet.cell(row = row_index, column = col_index).border = Border(top = Side(border_style='thin', 
																								color=colors.BLACK), 
																					 left  = Side(
																						 border_style='thin', 
																						 color=colors.BLACK), 
																					 right = Side(
																						 border_style='thin', 
																						 color=colors.BLACK), 
																					 bottom = Side(
																						 border_style='thin', 
																						 color=colors.BLACK))	
		
	except Exception as e:
		elog('border_line','{d}_{type}\n{trb1}\n{trb2} '.format(d = datetime.datetime.now(), 
																type = str(type(e)), 
																trb1 = traceback.format_tb(sys.exc_info()[2])[0], 
																trb2 = str(sys.exc_info()[1])))
		return []


def elog(name, msg): 
	
	''' creates log file with Args:
	name 	- string - name of log file
	msg 	- string - text of error message
	mem_loc - string - location in memory for saving log_file
	'''
	with open('{path}\\elog_{name}.txt'.format (name = name, path = path), "a") as log_file:
		log_file.write('\n{sep}'.format(sep = '--------------------------------------------------------------------'))
		log_file.write('\n{msg}'.format(msg = msg))
		log_file.write('\n{sep}'.format(sep = '--------------------------------------------------------------------'))

def find_corect_list(wb):
	
	#find correct list and initiate self.title
	# wb - loaded workbook
	title = ''
	try:
		str_check = 'report'
			
		for title in wb:
			if str_check in str(title): #check if including 'report' in name of Worksheet
				
				# cuting name of worksheet
				title = str(title)
				title = title.replace('<Worksheet "', '')
				title = title.replace('">', '')
		
		title = title	
		
	except	Exception as e:
		elog('find_corect_list','{d}_{type}\n{trb1}\n{trb2} '.format(d=datetime.datetime.now(), 
																	  type=str(type(e)), 
																	  trb1=traceback.format_tb(sys.exc_info()[2])[0], 
																	  trb2=str(sys.exc_info()[1])))
		return[]
	
	return title
	
		

def load_ohf(year, filename):
	
	try:
		date_start_year= datetime.date(year, 1	, 1 )
		date_end_year= datetime.date(year, 12	, 31)
		
		het_list = []
			
		het_nme_str			 = 'Entry Request Name'	.lower()
		stts_str = 'Status'				.lower()
		empl_nme_str = 'Member Name'			.lower()
		empl_id_str = 'ID'					.lower()
		empl_tem_str = 'Service Team'		.lower()
		strt_dte_str = 'Start Date'			.lower()
		end_dte_str = 'End Date'			.lower()
		entr_dte_str = 'Last Modified Date'	.lower()
		nts_str = 'Notes'				.lower()
		actvt_str = 'Activity'			.lower()
				
		wb = load_workbook(filename)
		sheet_ranges = wb[u'{x}'.format(x = find_corect_list (wb))]#initiating source sheet
		
		#looking for max row number
		for row_index in range(1,1000):
			if str(sheet_ranges.cell(row = row_index, column = 1).value).lower() == 'het_Russia'.lower():
				sht_max_row = row_index - 2
		
		for col_index in range(1,15):
			cur_cll = str(sheet_ranges.cell(row = 1, column = col_index).value).lower()
			if het_nme_str in cur_cll	:	het_nme_col = col_index
			if stts_str in cur_cll		:	stts_col = col_index
			if empl_nme_str in cur_cll	:	empl_nme_col = col_index
			if empl_id_str in cur_cll	:	empl_id_col = col_index
			if empl_tem_str in cur_cll	:	empl_tem_col = col_index
			if strt_dte_str in cur_cll	:	strt_dte_col = col_index
			if end_dte_str in cur_cll	:	end_dte_col = col_index
			if entr_dte_str in cur_cll	:	entr_dte_col = col_index
			if nts_str in cur_cll		:	nts_col = col_index
			if actvt_str in cur_cll		:	actvt_col = col_index
		
		for row_index in range(2, sht_max_row +1):
			#creating list_pay
			row_i = row_index
			if ((str(sheet_ranges.cell (row=row_i, column=het_nme_col ).value).lower() != 'Sick Time'.lower()) and
				(datetime.datetime.date(sheet_ranges.cell(row=row_i, column=strt_dte_col).value) >= date_start_year) and 
				(datetime.datetime.date(sheet_ranges.cell(row=row_i, column=end_dte_col).value) <= date_end_year) and
				(((sheet_ranges.cell(row=row_i, column=empl_tem_col).value).lower() == 'RU FS UPS'.lower()) or
				((sheet_ranges.cell(row=row_i, column=empl_tem_col).value).lower() == 'RU FS Cooling'.lower()) or
				((sheet_ranges.cell(row=row_i, column=empl_tem_col).value).lower() == 'RU FS UPS Regions'.lower()))):
			
				het_line={}
				
				het_line['het_nme']=str(sheet_ranges.cell(row=row_i,column=het_nme_col).value)
				het_line['stts']=str(sheet_ranges.cell(row=row_i,column=stts_col).value)
				het_line['empl_id']=str(sheet_ranges.cell(row=row_i,column=empl_id_col).value)
				het_line['empl_tm']=str(sheet_ranges.cell(row=row_i,column=empl_tem_col).value)
				het_line['nts']=str(sheet_ranges.cell(row=row_i,column=nts_col).value)
				het_line['strt_dte']=datetime.datetime.date(sheet_ranges.cell(row=row_i,column=strt_dte_col).value)
				het_line['end_dte']=datetime.datetime.date(sheet_ranges.cell(row=row_i,column=end_dte_col).value)
				het_line['drtn']=(het_line['end_dte']-het_line['strt_dte']+timedelta(days=1)).days
				het_line['entr_dte']=datetime.datetime.date(sheet_ranges.cell(row=row_i, column=entr_dte_col).value)
				
				if (het_line['nts'] == 'None') or (het_line['nts'] == '-'):
					het_line['nts'] = u'Без комментария'
				het_list.append(het_line)

		return het_list

	except Exception as e:
		elog('load_ohf','{d}_{type}\n{trb1}\n{trb2} '.format(d = datetime.datetime.now(),
															  type = str(type(e)),
															  trb1 = traceback.format_tb(sys.exc_info()[2])[0],
															  trb2 = str(sys.exc_info()[1])))
		return []

		

def make_objects(year, het_list):

	try:
	
		nme_sub_lst = []
		het_lst_by_nme = []
		
		for line_nme in nme_lst:
			for line_ohf in het_list:
				if str(line_nme['persone_id']).lower() == str(line_ohf['empl_id']).lower() :
					nme_sub_lst.append(str(line_nme['persone_id']).lower())
					
		nme_sub_lst = set(nme_sub_lst)
		
		#Combine by persone name
		for name in nme_sub_lst:
			
			for line_nme in nme_lst:
				if line_nme['persone_id'].lower() == name.lower():
				
					empl_line = {}
					
					empl_line['empl_id']= line_nme['persone_id']
					empl_line['empl_en_shrt_nme']= line_nme['persone_en_shrt_nme']
					empl_line['empl_ru_shrt_nme']= line_nme['persone_ru_shrt_nme']
					empl_line['empl_ru_lng_nme']= line_nme['persone_ru_lng_nme']
					empl_line['empl_srs_lctn']= line_nme['persone_srs_lctn']
					empl_line['tl_nme']= line_nme['tl_nme']
					empl_line['persone_tpe']= line_nme['persone_tpe']
					empl_line['persone_het_drtn']= 0
					empl_line['het_lst'] = []
					empl_line['het_days_lst'] = {}
					sub_het_lst = []
					
					for line_ohf in het_list:
						if str(line_ohf['empl_id']).lower() == name.lower() :
						
							empl_line['empl_tm']= line_ohf['empl_tm']
							empl_line['persone_het_drtn']= empl_line['persone_het_drtn'] + line_ohf['drtn']
							
							het_line = {}
							
							het_line['het_nme']= line_ohf['het_nme']
							het_line['strt_dte']= line_ohf['strt_dte']
							het_line['end_dte']= line_ohf['end_dte']
							het_line['entr_dte']= line_ohf['entr_dte']
							het_line['drtn']= line_ohf['drtn']
							het_line['stts']= line_ohf['stts']
							het_line['nts']= line_ohf['nts']
							het_line['days_lst']= []
							delta = het_line['end_dte'] - het_line['strt_dte']         # timedelta
							for i in range(delta.days + 1):
								day = het_line['strt_dte'] + timedelta(days=i)
								het_line['days_lst'].append(day) 
							
							sub_het_lst.append(het_line)
					
					sub_het_lst.sort(key = lambda item: (item['strt_dte']))
					
					indicator = 1
					new_sub_het_lst = []
					
					sub_het_lst_len = len(sub_het_lst)
						
					while indicator == 1 :
						indicator = 0
						i = 0
						new_sub_het_lst = []
						
						#Сombining in-coming het
						for line in sub_het_lst:
							if ((i < sub_het_lst_len - 1) and
							(sub_het_lst[i]['end_dte']+datetime.timedelta(days = 1)==sub_het_lst[i+1]['strt_dte']) and
							(sub_het_lst[i]['stts'] == sub_het_lst[i+1]['stts'])):
								new_line = []
								new_line = sub_het_lst[i]
								new_line['end_dte'] = sub_het_lst[i+1]['end_dte']
								new_line['drtn'] = sub_het_lst[i]['drtn'] + sub_het_lst[i+1]['drtn']
								new_line['days_lst']= sub_het_lst[i]['days_lst'] + sub_het_lst[i+1]['days_lst']
								new_sub_het_lst.append(new_line)
								indicator = 1
								i = i + 2
							elif (i < sub_het_lst_len):
								new_sub_het_lst.append(sub_het_lst[i])
								i = i + 1
						
						sub_het_lst = new_sub_het_lst
						sub_het_lst_len = len(new_sub_het_lst)
						
					empl_line['het_lst'] = sub_het_lst
					
					#Made empty days dir
					slovar = {}
					for i in range(0,367):
						if datetime.date(year, 1, 1) + timedelta(days = i) < datetime.date(year + 1, 1, 1):
							slovar['{day}'.format(day = datetime.date(year, 1, 1) + timedelta(days = i))] = 0
					empl_line['het_days_lst'] = slovar
					
					#Made filled days dir
					for line_t in empl_line['het_lst'] :
						for element_t in line_t['days_lst']:
							for line_d in empl_line['het_days_lst']:
								if str(element_t) == str(line_d) :
									empl_line['het_days_lst'][line_d] = empl_line['het_days_lst'][line_d] + 1

			het_lst_by_nme.append(empl_line)

		# het LIST BY DAY CREATIVE
		het_lst_by_day = []
	
		for i in range(0,367):
			if datetime.date(year, 1, 1) + timedelta(days = i) < datetime.date(year + 1, 1, 1):
				line = {}
				line['day'] = str(datetime.date(year, 1, 1) + timedelta(days = i))
				line['persone_lst'] = []
				line['persone_nme_lst'] = []
				het_lst_by_day.append(line)
		
		for line_n_1 in het_lst_by_nme:
			for line_n_2 in line_n_1['het_lst']:
				for element_n in line_n_2['days_lst']:
					for line_d in het_lst_by_day:
						if ((str(element_n) == str(line_d['day'])) and
							(line_n_1['empl_en_shrt_nme'] not in line_d['persone_nme_lst'])):
							persone_line = {}
							persone_line['empl_en_shrt_nme'] = line_n_1['empl_en_shrt_nme']
							persone_line['empl_ru_shrt_nme'] = line_n_1['empl_ru_shrt_nme']
							persone_line['empl_ru_lng_nme'] = line_n_1['empl_ru_lng_nme']
							persone_line['empl_srs_lctn'] = line_n_1['empl_srs_lctn']
							persone_line['tl_nme'] = line_n_1['tl_nme']
							persone_line['persone_tpe'] = line_n_1['persone_tpe']
							persone_line['het_nme'] = line_n_2['het_nme']
							persone_line['nts'] = line_n_2['nts']
							persone_line['strt_dte'] = line_n_2['strt_dte']
							persone_line['end_dte'] = line_n_2['end_dte']
							persone_line['entr_dte']= line_n_2['entr_dte']
							persone_line['drtn']= line_n_2['drtn']
							persone_line['stts']= line_n_2['stts']
							
							line_d['persone_lst'].append(persone_line)
							line_d['persone_nme_lst'].append(line_n_1['empl_en_shrt_nme'])
			
		return het_lst_by_nme, het_lst_by_day, nme_sub_lst
		
	except Exception as e:
		elog('make_objects','{d}_{type}\n{trb1}\n{trb2} '.format(d = datetime.datetime.now(),
																  type = str(type(e)),
																  trb1 = traceback.format_tb(sys.exc_info()[2])[0],
																  trb2 = str(sys.exc_info()[1])))
		return []



def find_error(het_lst_by_nme, het_lst_by_day, nme_sub_lst, year):
	
	errors = []
	try:
		#Check Traning&Vacation
		for line_t in trnng_lst_scnd:
			for persone in line_t['persone_lst']:
				for line_v in het_lst_by_nme:
					if persone['persone_id'].lower() == line_v['empl_id'].lower():
						for day in line_t['dys_lst']:
							if line_v['het_days_lst'][str(day)] > 0:
								a = u'У сотрудника {nme} на {dte} запланирован тренинг и отпуск.'.format(
									nme = line_v['empl_ru_shrt_nme'],
									dte = str(day))
								errors.append(a)
								
		#Check vacation existing
		for line_n in nme_lst:
			if str(line_n['persone_id']).lower() not in nme_sub_lst:
				a = u'Сотрудник {name} не начал планировать свой отпуск'.format(name = line_n['persone_ru_shrt_nme'])
				errors.append(a)
		
		for line in het_lst_by_nme:
			duration_dir = []
			for element in line['het_lst']:
				duration_dir.append(element['drtn'])
			
			#Check full duration
			if int(line['persone_het_drtn']) < 28 :
				a = u'Сотрудник {name} не до конца запланировал свой отпуск, запланировано дней: ' \
					u'{days}'.format(name = line['empl_ru_shrt_nme'], days = line['persone_het_drtn'])
				errors.append(a)
				
			#Check 2weeks duration
			if max(duration_dir) < 14 :
				a = u'Сотрудник {name} не имеет отпуск, продолжительность не менeе 2 недель, максимальный размер ' \
					u'непрерывного отпуска: {days} дней '.format(name = line['empl_ru_shrt_nme'],
																 days = max(duration_dir))
				errors.append(a)
				
			#Check maximum duration
			for het_line in line['het_lst']:
				if (het_line['drtn'] > 14) and (het_line['stts'].lower() != 'Approved'.lower()):
					a = u'Сотрудник {name} запланировал отпуск продолжительностью {days} дней. ' \
						u'Требует дополнительного согласования.'.format(
						name = line['empl_ru_shrt_nme'],
						days = het_line['drtn'])
					errors.append(a)
					
			#Check vacation after 01.10
			for element in line['het_lst']:
				if (element['strt_dte'] > datetime.date(year, 10, 1)) and (element['stts'].lower() != 'Approved'.lower()):
					a = u'Сотрудник {name} запланировал свой отпуск после 1 октября. ' \
						u'Требует дополнительного согласования.'.format(name = line['empl_ru_shrt_nme'])
					errors.append(a)
					
			#Check dublicate of het days
			for element in line['het_days_lst']:
				if line['het_days_lst'][element] >1:
					a = u'{name} допустил ошибку при формировании отпуска. ' \
						u'Более одного het создано на {day}.'.format(name = line['empl_ru_shrt_nme'], day = element)
					errors.append(a)

		fnctn_lst = [
		{'fnctn_nme' : 'SpetialCompany', 'persone':[]},
		{'fnctn_nme' : 'Cooling', 'persone':[]},
		{'fnctn_nme' : 'UPS', 'persone':[]},
]

		#Check persone in Region			
		for line_t in het_lst_by_day:
			
			lctn_el_lst= []	
			lctn_lst= []
			for line in nme_lst:
				lctn_el_lst.append(line['persone_srs_lctn'])
			lctn_el_lst = set(lctn_el_lst)
			for element in lctn_el_lst:
				line = {}
				line['lctn_nme'] = element
				line['persone'] = []
				lctn_lst.append(line)
			
			for het in line_t['persone_lst']:
				for lctn in lctn_lst:
					if (str(lctn['lctn_nme']).lower() == str(het['empl_srs_lctn']).lower()):
						line = {}
						line['persone_nme']= het['empl_ru_shrt_nme']
						line['persone_tpe']= het['persone_tpe']
						line['stts']= het['stts']
						line['het']= het['het_nme']
						line['entr_dte']= het['entr_dte']
						line['nts']= het['nts']
						
						lctn['persone'].append(line)

			for line in lctn_lst:
				cooling = 0
				ups = 0
				for persone in line['persone']:
					if persone['persone_tpe'].lower() == 'Cooling'.lower() :
						cooling += 1
				line['lnth_clng'] = cooling
				line['lnth_ups']= ups
			
			for lctn in lctn_lst:
				if (lctn['lctn_nme'].lower() != u'Москва'.lower()) and ((lctn['lnth_clng'] > 1) or (lctn['lnth_ups'] > 1)):
					a = u'{day} В городе {lctn} превышено допустимое значение инженеров, ' \
						u'одновременно находящихся в отпуске : '.format (lctn = lctn['lctn_nme'], day = line_t['day'])
					errors.append(a)
					for persone in lctn['persone']:
						a = u'{persone_nme}-{stts}-{het} {dte} - {nts}'.format(persone_nme = persone['persone_nme'],
																			   stts = persone['stts'],
																			   het = persone['het'],
																			   dte = persone['entr_dte'],
																			   nts = persone['nts'])
						errors.append(a)
					
		#Check persone in Moscow				
		for line_t in het_lst_by_day:
			fnctn_lst = [
				{	'fnctn_nme' : 'SpetialCompany'		,'persone'	:	[]		,'persone_len'	:	0	},
				{	'fnctn_nme' : 'Cooling'			,'persone'	:	[]		,'persone_len'	:	0	},
				{	'fnctn_nme' : 'UPS'				,'persone'	:	[]		,'persone_len'	:	0	},
]
				
			for het in line_t['persone_lst']:
				for line_f in fnctn_lst:
					if (het['empl_srs_lctn'].lower() == u'Москва'.lower()) and \
							(line_f['fnctn_nme'].lower() == het['persone_tpe'].lower()):
						line = {}
						line['persone_nme']= het['empl_ru_shrt_nme']
						line['persone_tpe']= het['persone_tpe']
						line['stts']= het['stts']
						line['het']= het['het_nme']
						line['entr_dte']= het['entr_dte']
						line['nts']= het['nts']
						
						line_f['persone'].append(line)
					
					line_f['persone_len'] = len(line_f['persone'])
			
			for fnctn in fnctn_lst:
				if (fnctn['fnctn_nme'].lower() == 'SpetialCompany'.lower()) and (len(fnctn['persone']) > 2) :
					a = u'{date} В Москве превышено допустимое значение инженеров, одновременно находящихся в отпуске' \
						u' по SpetialCompany :'.format(date = line_t['day'])
					errors.append(a)
					
					for persone in fnctn['persone']:
						a = u'{persone_nme}-{stts}-{het} {dte} - {nts}'.format(persone_nme = persone['persone_nme'],
																			   stts = persone['stts'],
																			   het = persone['het'],
																			   dte = persone['entr_dte'],
																			   nts = persone['nts'])
						errors.append(a)
						
				if (fnctn['fnctn_nme'].lower() == 'Cooling'.lower()) and (len(fnctn['persone']) > 2) :
					a = u'{date} В Москве превышено допустимое значение инженеров, одновременно находящихся ' \
						u'в отпуске по Cooling (возможно 2 человека, запланировали {persone_len} ' \
						u'человек):'.format(date = line_t['day'], persone_len = fnctn['persone_len'])
					errors.append(a)
					
					for persone in fnctn['persone']:
						a = u'{persone_nme}-{stts}-{het} {dte} - {nts}'.format(persone_nme = persone['persone_nme'],
																			   stts = persone['stts'],
																			   het = persone['het'],
																			   dte = persone['entr_dte'],
																			   nts = persone['nts'])
						errors.append(a)
						
				if (fnctn['fnctn_nme'].lower() == 'UPS'.lower()) and (len(fnctn['persone']) > 3) :
					a = u'{date} В Москве превышено допустимое значение инженеров, одновременно ' \
						u'находящихся в отпуске по UPS (возможно 3 человека, запланировали ' \
						u'{persone_len} человек):'.format(date = line_t['day'], persone_len = fnctn['persone_len'])
					errors.append(a)
					
					for persone in fnctn['persone']:
						a = u'{persone_nme}-{stts}-{het} {dte} - {nts}'.format(persone_nme = persone['persone_nme'],
																			   stts = persone['stts'],
																			   het = persone['het'],
																			   dte = persone['entr_dte'],
																			   nts = persone['nts'])
						errors.append(a)

		#Check persone by Function				
		for line_t in het_lst_by_day:
		
			fnctn_lst = [
				{	'fnctn_nme' : 'SpetialCompany'		,'persone'	:	[]		,'persone_len'	:	0	},
				{	'fnctn_nme' : 'Cooling'			,'persone'	:	[]		,'persone_len'	:	0	},
				{	'fnctn_nme' : 'UPS'				,'persone'	:	[]		,'persone_len'	:	0	},
]
		
			for het in line_t['persone_lst']:
				for line_f in fnctn_lst:
					if (line_f['fnctn_nme'] == het['persone_tpe']):
						line = {}
						line['persone_nme']= het['empl_ru_shrt_nme']
						line['persone_tpe']= het['persone_tpe']
						line['stts']= het['stts']
						line['het']= het['het_nme']
						line['entr_dte']= het['entr_dte']
						line['nts']= het['nts']
						
						line_f['persone'].append(line)
					
					line_f['persone_len'] = len(line_f['persone'])
			
			for fnctn in fnctn_lst:
				if (fnctn['fnctn_nme'].lower() == 'Cooling'.lower()) and (len(fnctn['persone']) > 3) :
					a = u'{date} Превышено допустимое значение инженеров, одновременно находящихся в отпуске по ' \
						u'Cooling (возможно 3 человека, запланировали {persone_len} человек):'.format(
						date = line_t['day'],
						persone_len = fnctn['persone_len'])
					errors.append(a)
					
					for persone in fnctn['persone']:
						a = u'{persone_nme}-{stts}-{het} {dte} - {nts}'.format(
							persone_nme = persone['persone_nme'],
							stts = persone['stts'],
							het = persone['het'],
							dte = persone['entr_dte'],
							nts = persone['nts'])
						errors.append(a)
						
				if (fnctn['fnctn_nme'].lower() == 'UPS'.lower()) and (len(fnctn['persone']) > 7) :
					a = u'{date} Превышено допустимое значение инженеров, одновременно находящихся в отпуске по UPS ' \
						u'(возможно 7 человек, запланировали {persone_len} человек):'.format(
						date = line_t['day'],
						persone_len = fnctn['persone_len'])
					errors.append(a)
					
					for persone in fnctn['persone']:
						a = u'{persone_nme}-{stts}-{het} {dte} - {nts}'.format(
							persone_nme = persone['persone_nme'],
							stts = persone['stts'],
							het = persone['het'],
							dte = persone['entr_dte'],
							nts = persone['nts'])
						errors.append(a)
						
		return 	errors, lctn_el_lst
		
	except Exception as e:
		elog('find_error','{d}_{type}\n{trb1}\n{trb2} '.format(
			d = datetime.datetime.now(),
			type = str(type(e)),
			trb1 = traceback.format_tb(sys.exc_info()[2])[0],
			trb2 = str(sys.exc_info()[1])))
		return []	

	

def make_chart(errors, lctn_el_lst) :
	
	try: 
		wb = load_workbook(filename = 'path')
		
		
		ws_a = wb["Comments"]
		ws_b = wb.create_sheet("Moscow")
		ws_c = wb.create_sheet("UPS")
		ws_d = wb.create_sheet("Cooling")
		ws_e= wb.create_sheet("Conditions")
		
		ws_b.freeze_panes = ws_b['B4']
		ws_c.freeze_panes = ws_c['B4']
		ws_d.freeze_panes = ws_d['B4']
		
		ws_a.sheet_view.showGridLines = False
		ws_b.sheet_view.showGridLines = False
		ws_c.sheet_view.showGridLines = False
		ws_d.sheet_view.showGridLines = False
		ws_e.sheet_view.showGridLines = False
		
		
		#------------------C O N D I T I O N S --------------------
		
		for col_index in range (1, 10):
			for row_index in range (1, 1000):
				ws_e.cell(row = row_index, column = col_index).value = ''	
		
		conditions = []
		
		i = 2
		for line in conditions:
			ws_e.cell(row = i, column = 2).value = line
			i = i + 1
		
		#------------------C O M M E N T S --------------------
		
		for col_index in range (1, 10):
			for row_index in range (1, 100):
				ws_a.cell(row = row_index, column = col_index).value = ''	
		
		ws_a.column_dimensions['A'].width = 3
		ws_a.column_dimensions['B'].width = 100
		
		i = 2
		
		for line in errors:
			ws_a.cell(row = i, column = 2).value = line
			i = i + 1

		#------------------M O S C O W --------------------
		
		ws_b.column_dimensions['A'].width = 10
			
		for i in range(3, 3 + len(het_lst_by_day)) :
			ws_b.row_dimensions[i].height = 15
		
		clmn_len = 2
		for nme in het_lst_by_nme :
			if  nme['empl_srs_lctn'].lower() == u'Москва'.lower() :
				ws_b.column_dimensions[get_column_letter(clmn_len)].width = 9
				clmn_len = clmn_len + 1
		
		fnt_1 = Font(name='Arial', size=10)
		fnt_2 = Font(name='Arial', size=10)
		fnt_3 = Font(name='Arial', size=8)
		fnt_4 = Font(name='Arial', size=8, bold = True)
		fll_1 = PatternFill(start_color = '70D248', end_color = '70D248', fill_type = "solid")
		fll_2 = PatternFill(start_color = 'F0D94D', end_color = 'F0D94D', fill_type = "solid")
		fll_3 = PatternFill(start_color = 'A2D4E2', end_color = 'A2D4E2', fill_type = "solid")
		
		ws_b['A1'].value = u"Function"
		ws_b['A2'].value = u"persone"
		ws_b['A3'].value = u"id"
		for i in range(4, 4 + len(het_lst_by_day)) :
			ws_b.cell(row = i, column = 1).value = het_lst_by_day[i-4]['day']
			a= ws_b.cell(row = i, column = 1)
			a.font = fnt_1
		
		lst_clmn = 1
		
		frst_clmn = lst_clmn + 1
		for nme in het_lst_by_nme :
			if ((nme['empl_srs_lctn'].lower() == u'Москва'.lower()) and (nme['persone_tpe'].lower() == 'UPS'.lower())):
				ws_b.cell(row = 2, column = lst_clmn + 1).value = nme['empl_en_shrt_nme'].title()
				ws_b.cell(row = 3, column = lst_clmn + 1).value = nme['empl_id'].lower()
				a = ws_b.cell(row = 2, column = lst_clmn + 1 )
				b = ws_b.cell(row = 3, column = lst_clmn + 1 )
				a.alignment = Alignment(text_rotation=90, horizontal='center')
				b.alignment = Alignment(horizontal='center')
				a.font = fnt_2
				b.font = fnt_3
				lst_clmn = lst_clmn + 1
		
		for row_index in range(1, len(het_lst_by_day) + 4):
			ws_b.cell(row = row_index, column = frst_clmn).border = Border(right = Side(
				border_style='double',
				color=colors.BLACK))
				
		if frst_clmn != lst_clmn :
			ws_b.merge_cells('{strt_col}1:{end_col}1'.format (strt_col = get_column_letter(frst_clmn),
															  end_col = get_column_letter(lst_clmn)))
			ws_b['{strt_col}1'.format (strt_col = get_column_letter(frst_clmn))].value = "UPS"
			
			
		frst_clmn = lst_clmn + 1
		for nme in het_lst_by_nme :
			if ((nme['empl_srs_lctn'].lower() == u'Москва'.lower()) and (nme['persone_tpe'].lower() == 'Cooling'.lower())):
				ws_b.cell(row = 2, column = lst_clmn + 1).value = nme['empl_en_shrt_nme'].title()
				ws_b.cell(row = 3, column = lst_clmn + 1).value = nme['empl_id'].lower()
				a = ws_b.cell(row = 2, column = lst_clmn + 1 )
				b = ws_b.cell(row = 3, column = lst_clmn + 1 )
				a.alignment = Alignment(text_rotation=90, horizontal='center')
				b.alignment = Alignment(horizontal='center')
				a.font = fnt_2
				b.font = fnt_3
				lst_clmn = lst_clmn + 1
		if frst_clmn != lst_clmn :
			ws_b.merge_cells('{strt_col}1:{end_col}1'.format (strt_col=get_column_letter(frst_clmn),
															  end_col=get_column_letter(lst_clmn)))
			ws_b['{strt_col}1'.format (strt_col = get_column_letter(frst_clmn))].value = "Cooling"
		
		frst_clmn = lst_clmn + 1
		for nme in het_lst_by_nme :
			if ((nme['empl_srs_lctn'].lower() == u'Москва'.lower()) and (nme['persone_tpe'].lower() == 'SpetialCompany'.lower())):
				ws_b.cell(row = 2, column = lst_clmn + 1).value = nme['empl_en_shrt_nme'].title()
				ws_b.cell(row = 3, column = lst_clmn + 1).value = nme['empl_id'].lower()
				a = ws_b.cell(row = 2, column = lst_clmn + 1 )
				b = ws_b.cell(row = 3, column = lst_clmn + 1 )
				a.alignment =Alignment(text_rotation=90, horizontal='center')
				b.alignment =Alignment(horizontal='center')
				a.font = fnt_2
				b.font = fnt_3
				lst_clmn = lst_clmn + 1
		
		if frst_clmn != lst_clmn + 1 :
			ws_b.merge_cells('{strt_col}1:{end_col}1'.format (strt_col = get_column_letter(frst_clmn),
															  end_col = get_column_letter(lst_clmn)))
			ws_b['{strt_col}1'.format (strt_col = get_column_letter(frst_clmn))].value = "SpetialCompany"
		
		#Vacation
		i = 4
		for line in het_lst_by_day:
			day = line['day']
			for persone in line['persone_lst']:
				for j in range(2,lst_clmn + 1):
					if ((str(persone['empl_en_shrt_nme']).lower() == str(ws_b.cell(column=j, row=2).value).lower()) and
						(str(persone['stts']).lower() == 'Approved'.lower())):
							ws_b.cell(column = j, row = i).fill = fll_1
							ws_b.cell(column = j, row = i).font= fnt_3
							ws_b.cell(column = j, row = i).value= persone['entr_dte']
					elif ((str(persone['empl_en_shrt_nme']).lower() == str(ws_b.cell(column=j, row=2).value).lower()) and
						(str(persone['stts']).lower() != 'Approved'.lower())):
							ws_b.cell(column = j, row = i).fill= fll_2
							ws_b.cell(column = j, row = i).font= fnt_3
							ws_b.cell(column = j, row = i).value= persone['entr_dte']
			i = i + 1
		
		#Training
		for line in trnng_lst_scnd:
			for day in line['dys_lst']:
				for i in range(3, 3 + len(het_lst_by_day)):
					if (str(day) == ws_b.cell(column = 1, row = i).value) :
						row_index = i
				
				for persone in line['persone_lst']:
					for j in range(2,lst_clmn + 1):
						if (str(persone['persone_nme']).lower() == str(ws_b.cell(column = j, row = 2).value).lower()) :
							ws_b.cell(column = j, row = row_index).value= 'Training'
							ws_b.cell(column = j, row = row_index).font= fnt_4
		
		#Weekends
		for row_index in range(4, 4 + len(het_lst_by_day)) :
			for day_c in calendar:
				if (str(ws_b.cell(column = 1, row = row_index).value) == day_c) and (calendar[day_c] == 'n'):
					ws_b.cell(column = 1, row = row_index).fill = fll_3
		
		border_line(ws_b,1, 4 + len(het_lst_by_day),1,clmn_len)
		
		#------------------U P S--------------------	
		
		ws_c.column_dimensions['A'].width = 10
			
		for i in range(4, 4 + len(het_lst_by_day)) :
			ws_c.row_dimensions[i].height = 15
		
		clmn_len = 2
		for nme in het_lst_by_nme :
			if  nme['persone_tpe'].lower() == u'UPS'.lower() :
				ws_c.column_dimensions[get_column_letter(clmn_len)].width = 9
				clmn_len = clmn_len + 1
		
		ws_c['A1'].value = u"Region"
		ws_c['A2'].value = u"persone"
		ws_c['A3'].value = u"id"
		for i in range(4, 4 + len(het_lst_by_day)) :
			ws_c.cell(row = i, column = 1).value = het_lst_by_day[i-4]['day']
			a = ws_c.cell(row = i, column = 1)
			a.font = fnt_1
		
		lst_clmn = 1
		
		frst_clmn = lst_clmn + 1
		for lctn in lctn_el_lst:
			frst_clmn = lst_clmn + 1
			for nme in het_lst_by_nme :
				if ((nme['empl_srs_lctn'].lower() == lctn.lower()) and (nme['persone_tpe'].lower() == 'UPS'.lower())):
					ws_c.cell(row = 2, column = lst_clmn + 1).value = nme['empl_en_shrt_nme'].title()
					ws_c.cell(row = 3, column = lst_clmn + 1).value = nme['empl_id'].lower()
					a = ws_c.cell(row = 2, column = lst_clmn + 1 )
					b = ws_c.cell(row = 3, column = lst_clmn + 1 )
					a.alignment = Alignment(text_rotation=90, horizontal='center')
					b.alignment = Alignment(horizontal='center')
					a.font = fnt_2
					b.font = fnt_3
					lst_clmn = lst_clmn + 1
		
			if frst_clmn != lst_clmn + 1:
				ws_c.merge_cells('{strt_col}1:{end_col}1'.format (strt_col = get_column_letter(frst_clmn),
																  end_col = get_column_letter(lst_clmn)))
				ws_c['{strt_col}1'.format (strt_col = get_column_letter(frst_clmn))].value = lctn
	
		i = 4
		#Vacation
		for line in het_lst_by_day:
			day = line['day']
			for persone in line['persone_lst']:
				for j in range(2,lst_clmn + 1):
					if ((str(persone['empl_en_shrt_nme']).lower() == str(ws_c.cell(column = j, row = 2).value).lower()) and 
						(str(persone['stts']).lower() == 'Approved'.lower())):
							ws_c.cell(column = j, row = i).fill = fll_1
							ws_c.cell(column = j, row = i).font= fnt_3
							ws_c.cell(column = j, row = i).value= persone['entr_dte']
					elif ((str(persone['empl_en_shrt_nme']).lower() == str(ws_c.cell(column = j, row = 2).value).lower()) and 
						(str(persone['stts']).lower() != 'Approved'.lower())):
							ws_c.cell(column = j, row = i).fill = fll_2
							ws_c.cell(column = j, row = i).font= fnt_3
							ws_c.cell(column = j, row = i).value= persone['entr_dte']
			i = i + 1
		
		#Training
		for line in trnng_lst_scnd:
			for day in line['dys_lst']:
				for i in range(3, 3 + len(het_lst_by_day)):
					if (str(day) == ws_c.cell(column = 1, row = i).value) :
						row_index = i
				
				for persone in line['persone_lst']:
					for j in range(2,lst_clmn + 1):
						if (str(persone['persone_nme']).lower() == str(ws_c.cell(column = j, row = 2).value).lower()) :
							ws_c.cell(column = j, row = row_index).value= 'Training'
							ws_c.cell(column = j, row = row_index).font= fnt_4
		
		#Weekends
		for row_index in range(4, 4 + len(het_lst_by_day)) :
			for day_c in calendar:
				if (str(ws_c.cell(column = 1, row = row_index).value) == day_c) and (calendar[day_c] == 'n'):
					ws_c.cell(column = 1, row = row_index).fill = fll_3
		
		border_line(ws_c,1,4 + len(het_lst_by_day),1,clmn_len)
			
		#-----------------C O O L I N G--------------------	
		
		ws_d.column_dimensions['A'].width = 10
			
		for i in range(4, 4 + len(het_lst_by_day)) :
			ws_d.row_dimensions[i].height = 15
		
		clmn_len = 2
		for nme in het_lst_by_nme :
			if  nme['persone_tpe'].lower() == u'Cooling'.lower() :
				ws_d.column_dimensions[get_column_letter(clmn_len)].width = 9
				clmn_len = clmn_len + 1
		

		ws_d['A1'].value = u"Region"
		ws_d['A2'].value = u"persone"
		ws_d['A3'].value = u"id"
		for i in range(4, 4 + len(het_lst_by_day)) :
			ws_d.cell(row = i, column = 1).value = het_lst_by_day[i-4]['day']
			a = ws_d.cell(row = i, column = 1)
			a.font = fnt_1
		
		lst_clmn = 1
		
		frst_clmn = lst_clmn + 1
		for lctn in lctn_el_lst:
			frst_clmn = lst_clmn + 1
			for nme in het_lst_by_nme :
				if ((nme['empl_srs_lctn'].lower() == lctn.lower()) and (nme['persone_tpe'].lower() == 'Cooling'.lower())):
					ws_d.cell(row = 2, column = lst_clmn + 1).value = nme['empl_en_shrt_nme'].title()
					ws_d.cell(row = 3, column = lst_clmn + 1).value = nme['empl_id'].lower()
					a = ws_d.cell(row = 2, column = lst_clmn + 1 )
					b = ws_d.cell(row = 3, column = lst_clmn + 1 )
					a.alignment =Alignment(text_rotation=90, horizontal='center')
					b.alignment =Alignment(horizontal='center')
					a.font = fnt_2
					b.font = fnt_3
					lst_clmn = lst_clmn + 1
		
			if frst_clmn != lst_clmn + 1:
				ws_d.merge_cells('{strt_col}1:{end_col}1'.format (strt_col = get_column_letter(frst_clmn), end_col = get_column_letter(lst_clmn)))
				ws_d['{strt_col}1'.format (strt_col = get_column_letter(frst_clmn))].value = lctn
	
		i = 4
		for line in het_lst_by_day:
			day = line['day']
			for persone in line['persone_lst']:
				for j in range(2,lst_clmn + 1):
					if ((str(persone['empl_en_shrt_nme']).lower() == str(ws_d.cell(column = j, row = 2).value).lower()) and 
						(str(persone['stts']).lower() == 'Approved'.lower())):
							ws_d.cell(column = j, row = i).fill = fll_1
							ws_d.cell(column = j, row = i).font= fnt_3
							ws_d.cell(column = j, row = i).value= persone['entr_dte']
					elif ((str(persone['empl_en_shrt_nme']).lower() == str(ws_d.cell(column = j, row = 2).value).lower()) and 
						(str(persone['stts']).lower() != 'Approved'.lower())):
							ws_d.cell(column = j, row = i).fill = fll_2
							ws_d.cell(column = j, row = i).font= fnt_3
							ws_d.cell(column = j, row = i).value= persone['entr_dte']
			i = i + 1
		
		#Training
		for line in trnng_lst_scnd:
			for day in line['dys_lst']:
				for i in range(3, 3 + len(het_lst_by_day)):
					if (str(day) == ws_d.cell(column = 1, row = i).value) :
						row_index = i
				
				for persone in line['persone_lst']:
					for j in range(2,lst_clmn + 1):
						if (str(persone['persone_nme']).lower() == str(ws_d.cell(column = j, row = 2).value).lower()):
							ws_d.cell(column = j, row = row_index).value= 'Training'
							ws_d.cell(column = j, row = row_index).font= fnt_4

		
		#Weekends
		for row_index in range(4, 4 + len(het_lst_by_day)) :
			for day_c in calendar:
				if (str(ws_d.cell(column = 1, row = row_index).value) == day_c) and (calendar[day_c] == 'n'):
					ws_d.cell(column = 1, row = row_index).fill = fll_3
		
		border_line(ws_d,1,4 + len(het_lst_by_day),1,clmn_len)
		
		now = datetime.datetime.now()
		date = '{y}.{m}.{d}'.format(y = now.year, m = now.month, d = now.day)
		time = '{h}.{m}.{s}'.format(h = now.hour, m = now.minute, s = now.second)
		wb.save('{path}\\Vacation_{date}_{time}.xlsx'.format(path = path, date = date, time = time))
	
	except Exception as e:
		elog('make_chart','{d}_{type}\n{trb1}\n{trb2} '.format(d = datetime.datetime.now(),
																type = str(type(e)),
																trb1 = traceback.format_tb(sys.exc_info()[2])[0],
																trb2 = str(sys.exc_info()[1])))
		return []
	
		
year = 2020
het_list = load_ohf(year , 'path')
het_lst_by_nme, het_lst_by_day, nme_sub_lst = make_objects(year, het_list)
errors, lctn_el_lst = find_error(het_lst_by_nme, het_lst_by_day, nme_sub_lst, year)
make_chart(errors, lctn_el_lst)
		